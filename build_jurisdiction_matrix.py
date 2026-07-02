"""Parse the KYC.com Jurisdiction Matrix workbook into a flat JSON the app can
serve with zero runtime dependencies.

Source of truth: the skill's reference copy
  C:\\Users\\unzer\\.claude\\skills\\knowyourcustomer\\references\\JurisdictionMatrix.xlsx

Run this whenever KYC.com ships an updated matrix (coverage is fluid):
  python build_jurisdiction_matrix.py
→ writes jurisdiction_matrix.json next to this script.

Layout of each sheet: every jurisdiction is a VERTICAL block of merged rows.
Col A = jurisdiction name (only on the block's first row; blank-A rows continue
the block). Field/document lists run DOWN columns D-G across the block's rows.
Column G interleaves 'Mandatory:' / 'Non-Mandatory:' sub-headers with doc names.
"""
import os, json, sys
import openpyxl

SRC = r"C:\Users\unzer\.claude\skills\knowyourcustomer\references\JurisdictionMatrix.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jurisdiction_matrix.json")

# column indexes (0-based) within a row tuple
C_NAME, C_ACCESS, C_SLA, C_IDENTITY, C_CONTROL, C_SHARE, C_DOCS = 0, 1, 2, 3, 4, 5, 6


def _clean(v):
    return "" if v is None else str(v).strip()


def _is_na(v):
    return v.strip().lower() in ("", "n/a", "na", "-", "–")


def parse_block(rows):
    """rows = list of row-tuples belonging to ONE jurisdiction block."""
    name = _clean(rows[0][C_NAME])
    access = ""
    sla = ""
    identity, control, share = [], [], []
    docs_m, docs_n = [], []
    section = "mandatory"  # column G starts under Mandatory by convention
    for r in rows:
        if not access:
            access = _clean(r[C_ACCESS])
        if not sla:
            sla = _clean(r[C_SLA])
        vi, vc, vs = _clean(r[C_IDENTITY]), _clean(r[C_CONTROL]), _clean(r[C_SHARE])
        if vi and not _is_na(vi):
            identity.append(vi)
        if vc and not _is_na(vc):
            control.append(vc)
        if vs and not _is_na(vs):
            share.append(vs)
        vg = _clean(r[C_DOCS])
        if vg:
            low = vg.lower().rstrip(":")
            if low.startswith("mandatory") and "non" not in low:
                section = "mandatory"
            elif low.startswith("non-mandatory") or low.startswith("non mandatory") or low.startswith("nonmandatory"):
                section = "non_mandatory"
            elif not _is_na(vg):
                (docs_m if section == "mandatory" else docs_n).append(vg)
    return {
        "name": name,
        "registryAccess": access,
        "sla": sla,
        "companyIdentity": identity,
        "controlling": control,
        "shareholders": share,          # [] means registry returns no shareholder layer (n/a)
        "documentsMandatory": docs_m,
        "documentsNonMandatory": docs_n,
    }


def parse_sheet(ws, sheet_name):
    out = []
    block = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # pad short rows
        row = tuple((row + (None,) * 7)[:7])
        if _clean(row[C_NAME]):
            if block:
                out.append(parse_block(block))
            block = [row]
        else:
            if block:
                block.append(row)
    if block:
        out.append(parse_block(block))
    for j in out:
        j["group"] = sheet_name
    return out


def main():
    if not os.path.exists(SRC):
        sys.exit("Matrix workbook not found: " + SRC)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    groups = []
    all_j = []
    for sn in wb.sheetnames:
        js = parse_sheet(wb[sn], sn)
        groups.append({"name": sn, "count": len(js)})
        all_j.extend(js)
    data = {
        "source": "Jurisdiction coverage matrix",
        "note": ("Coverage is fluid and extended into new jurisdictions over time. "
                 "Documents are split into Base (provided as standard) and Additional "
                 "(available on request, not commonly needed, billed as an extra charge) "
                 "so onboarding stays cost-effective."),
        "groups": groups,
        "total": len(all_j),
        "jurisdictions": all_j,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"Wrote {OUT}: {len(all_j)} jurisdictions across {len(groups)} groups")
    for g in groups:
        print(f"  - {g['name']}: {g['count']}")


if __name__ == "__main__":
    main()
